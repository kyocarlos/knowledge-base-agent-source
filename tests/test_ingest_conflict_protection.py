from pathlib import Path

import pytest
from openpyxl import Workbook

from src.ingest_conflict_protection import (
    IngestContractError,
    build_document_id,
    build_idempotency_key,
    read_km_metadata,
    validate_ingest_file,
)
from src.ingest_registry import IngestRegistry, IngestRegistryConflict


def _identity() -> dict[str, str]:
    return {
        "sourceSystem": "anritsu",
        "environmentId": "env-local",
        "projectId": "project-1",
        "runId": "run-001",
        "artifactType": "single",
        "reportSchema": "1.0",
        "originalFileName": "report.xlsx",
        "sourceFileHash": "a" * 64,
        "generatedAt": "2026-08-03T00:00:00Z",
    }


def test_identity_derivation_is_deterministic():
    metadata = _identity()
    metadata["documentId"] = build_document_id("4g5g", metadata)
    metadata["idempotencyKey"] = build_idempotency_key(metadata)
    assert metadata["documentId"] == "4g5g:anritsu:env-local:run-001:single"
    assert len(metadata["idempotencyKey"]) == 64
    assert build_idempotency_key(metadata) == metadata["idempotencyKey"]


def test_registry_deduplicates_and_rejects_document_overwrite(tmp_path: Path):
    registry = IngestRegistry(tmp_path / "registry.sqlite3")
    metadata = _identity()
    metadata["documentId"] = build_document_id("4g5g", metadata)
    metadata["ingestFileHash"] = "b" * 64
    metadata["idempotencyKey"] = build_idempotency_key(metadata)
    identity = {
        "source_system": metadata["sourceSystem"],
        "environment_id": metadata["environmentId"],
        "project_id": metadata["projectId"],
        "run_id": metadata["runId"],
        "artifact_type": metadata["artifactType"],
        "report_schema": metadata["reportSchema"],
        "original_file_name": metadata["originalFileName"],
        "source_file_hash": metadata["sourceFileHash"],
        "ingest_file_hash": metadata["ingestFileHash"],
        "document_id": metadata["documentId"],
        "idempotency_key": metadata["idempotencyKey"],
        "generated_at": metadata["generatedAt"],
    }

    first, duplicate = registry.register(identity, "task-1")
    assert duplicate is False
    second, duplicate = registry.register(identity, "task-2")
    assert duplicate is True
    assert second["task_id"] == first["task_id"] == "task-1"

    different = dict(identity, idempotency_key="c" * 64)
    with pytest.raises(IngestRegistryConflict) as conflict:
        registry.register(different, "task-3")
    assert conflict.value.code == "document_conflict"


def test_km_metadata_sheet_is_read_without_filename_assumptions(tmp_path: Path):
    path = tmp_path / "report.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KM_Metadata"
    sheet.append(["key", "value"])
    sheet.append(["sourceSystem", "amarisoft"])
    sheet.append(["runId", "run-2"])
    workbook.save(path)
    assert read_km_metadata(path)["sourceSystem"] == "amarisoft"


def test_missing_identity_headers_are_explicit(tmp_path: Path):
    path = tmp_path / "report.xlsx"
    workbook = Workbook()
    workbook.active.title = "KM_Metadata"
    workbook.save(path)
    with pytest.raises(IngestContractError) as error:
        validate_ingest_file(path=path, headers={}, extraction_mode="4g5g")
    assert error.value.code == "headers_missing"


def test_ingest_hash_is_server_generated_without_circular_metadata(tmp_path: Path):
    path = tmp_path / "report.xlsx"
    metadata = _identity()
    metadata["documentId"] = build_document_id("4g5g", metadata)
    metadata["idempotencyKey"] = build_idempotency_key(metadata)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "KM_Metadata"
    sheet.append(["key", "value"])
    for key, value in metadata.items():
        sheet.append([key, value])
    workbook.save(path)
    headers = {
        "Authorization": "Bearer test",
        "Idempotency-Key": metadata["idempotencyKey"],
        "X-KB-Source-System": metadata["sourceSystem"],
        "X-KB-Environment-Id": metadata["environmentId"],
        "X-KB-Run-Id": metadata["runId"],
        "X-KB-Artifact-Type": metadata["artifactType"],
        "X-KB-Document-Id": metadata["documentId"],
    }
    identity = validate_ingest_file(path=path, headers=headers, extraction_mode="4g5g", original_file_name="ANRITSU__envelope__report.xlsx")
    assert len(identity.ingest_file_hash) == 64
