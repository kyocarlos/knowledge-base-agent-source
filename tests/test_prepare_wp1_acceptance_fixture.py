from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from prepare_wp1_acceptance_fixture import (
    ISOLATED_RUN_PREFIX,
    RUN_PREFIX,
    build_fixture,
    fixture_run_id,
    request_contract,
    validate_request_contract,
)


def test_generated_fixture_and_request_contract_share_one_run_id(tmp_path):
    result = build_fixture(tmp_path, RUN_PREFIX + "fixed-test-id")
    assert fixture_run_id(tmp_path / "TR-E2E-WP1-PROD-fixed-test-id.xlsx") == result["run_id"]
    contract = request_contract(result["run_id"])
    assert contract["X-E2E-Test-Run-ID"] == result["run_id"]
    assert contract["Idempotency-Key"] == result["run_id"]
    assert contract["cleanup_test_run_id"] == result["run_id"]
    assert validate_request_contract(tmp_path / "TR-E2E-WP1-PROD-fixed-test-id.xlsx", result["run_id"]) == contract


def test_stale_fixture_is_rejected_by_explicit_manifest_check(tmp_path):
    result = build_fixture(tmp_path, RUN_PREFIX + "manifest-id")
    with pytest.raises(ValueError, match="must be identical"):
        import openpyxl

        workbook = openpyxl.load_workbook(result["fixture"])
        sheet = workbook["Manifest"]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == "test_run_id":
                sheet.cell(row, 2).value = RUN_PREFIX + "different-id"
        workbook.save(result["fixture"])
        fixture_run_id(tmp_path / "TR-E2E-WP1-PROD-manifest-id.xlsx")


def test_request_run_id_mismatch_is_rejected_before_network_call(tmp_path):
    result = build_fixture(tmp_path, RUN_PREFIX + "manifest-id")
    with pytest.raises(ValueError, match="must equal Manifest run_id"):
        validate_request_contract(Path(result["fixture"]), RUN_PREFIX + "request-id")


def test_missing_required_attachment_is_rejected_before_network_call(tmp_path):
    result = build_fixture(tmp_path, RUN_PREFIX + "attachment-id")
    Path(result["attachment"]).unlink()
    with pytest.raises(ValueError, match="attachment is missing"):
        validate_request_contract(Path(result["fixture"]), result["run_id"])


def test_each_generated_run_id_is_fresh(tmp_path):
    first = build_fixture(tmp_path / "one")
    second = build_fixture(tmp_path / "two")
    assert first["run_id"] != second["run_id"]


def test_isolated_fixture_uses_isolated_run_id_contract(tmp_path):
    result = build_fixture(tmp_path, ISOLATED_RUN_PREFIX + "fixed-test-id", "isolated")
    assert fixture_run_id(Path(result["fixture"])) == result["run_id"]
    assert validate_request_contract(Path(result["fixture"]), result["run_id"])["X-E2E-Test-Run-ID"] == result["run_id"]


def test_production_fixture_rejects_isolated_run_id(tmp_path):
    with pytest.raises(ValueError, match="production E2E format"):
        build_fixture(tmp_path, ISOLATED_RUN_PREFIX + "wrong-mode", "production")


def test_isolated_fixture_rejects_production_run_id(tmp_path):
    with pytest.raises(ValueError, match="isolated E2E format"):
        build_fixture(tmp_path, RUN_PREFIX + "wrong-mode", "isolated")
