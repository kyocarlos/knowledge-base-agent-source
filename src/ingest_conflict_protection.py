"""Conflict-protection contract for external knowledge-base ingestion."""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_HEADERS = (
    "authorization",
    "idempotency-key",
    "x-kb-source-system",
    "x-kb-environment-id",
    "x-kb-run-id",
    "x-kb-artifact-type",
    "x-kb-document-id",
)
REQUIRED_METADATA = (
    "sourceSystem",
    "environmentId",
    "projectId",
    "runId",
    "artifactType",
    "reportSchema",
    "originalFileName",
    "sourceFileHash",
    "documentId",
    "idempotencyKey",
    "generatedAt",
)
_SAFE_PART = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._:-]{0,127}$")
_ALIASES = {
    key.lower(): key
    for key in REQUIRED_METADATA
} | {
    "source_system": "sourceSystem",
    "environment_id": "environmentId",
    "project_id": "projectId",
    "run_id": "runId",
    "artifact_type": "artifactType",
    "report_schema": "reportSchema",
    "original_file_name": "originalFileName",
    "source_file_hash": "sourceFileHash",
    "ingest_file_hash": "ingestFileHash",
    "document_id": "documentId",
    "idempotency_key": "idempotencyKey",
    "generated_at": "generatedAt",
}


class IngestContractError(ValueError):
    def __init__(self, code: str, message: str, *, status_code: int = 422, fields: list[str] | None = None):
        self.code = code
        self.status_code = status_code
        self.fields = fields or []
        super().__init__(message)


@dataclass(frozen=True)
class IngestIdentity:
    source_system: str
    environment_id: str
    project_id: str
    run_id: str
    artifact_type: str
    report_schema: str
    original_file_name: str
    source_file_hash: str
    ingest_file_hash: str
    document_id: str
    idempotency_key: str
    generated_at: str

    def as_dict(self) -> dict[str, str]:
        return {
            "source_system": self.source_system,
            "environment_id": self.environment_id,
            "project_id": self.project_id,
            "run_id": self.run_id,
            "artifact_type": self.artifact_type,
            "report_schema": self.report_schema,
            "original_file_name": self.original_file_name,
            "source_file_hash": self.source_file_hash,
            "ingest_file_hash": self.ingest_file_hash,
            "document_id": self.document_id,
            "idempotency_key": self.idempotency_key,
            "generated_at": self.generated_at,
        }


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _metadata_key(value: Any) -> str:
    raw = _text(value).replace(" ", "_").replace("-", "_").lower()
    return _ALIASES.get(raw, "")


def read_km_metadata(path: str | Path) -> dict[str, str]:
    """Read KM_Metadata in either key/value or one-row table form."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise IngestContractError("metadata_invalid", f"無法讀取 Excel KM_Metadata: {exc}") from exc
    try:
        sheet = next((item for item in workbook.worksheets if item.title.strip().lower() == "km_metadata"), None)
        if sheet is None:
            raise IngestContractError("metadata_missing", "Excel 缺少 KM_Metadata 工作表")
        rows = [[_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        rows = [row for row in rows if any(row)]
        if not rows:
            return {}
        result: dict[str, str] = {}
        first_keys = [_metadata_key(value) for value in rows[0]]
        if "sourceSystem" in first_keys and len(rows) > 1:
            values = rows[1]
            result.update({key: values[index] for index, key in enumerate(first_keys) if key and index < len(values)})
        else:
            for row in rows:
                if len(row) >= 2:
                    key = _metadata_key(row[0])
                    if key:
                        result[key] = row[1]
        return {key: _text(value) for key, value in result.items()}
    finally:
        workbook.close()


def _required(value: str, field: str) -> str:
    if not value:
        raise IngestContractError("metadata_missing", f"KM_Metadata 缺少必要欄位: {field}", fields=[field])
    if field not in {"originalFileName", "reportSchema", "generatedAt"} and not _SAFE_PART.fullmatch(value):
        raise IngestContractError("metadata_invalid", f"欄位格式不合法: {field}", fields=[field])
    return value


def build_idempotency_key(metadata: dict[str, str]) -> str:
    source = "\n".join(metadata[field] for field in ("sourceSystem", "environmentId", "runId", "artifactType", "sourceFileHash"))
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


def build_document_id(extraction_mode: str, metadata: dict[str, str]) -> str:
    return ":".join((extraction_mode, metadata["sourceSystem"], metadata["environmentId"], metadata["runId"], metadata["artifactType"]))


def validate_ingest_file(
    *,
    path: str | Path,
    headers: Any,
    extraction_mode: str,
    original_file_name: str | None = None,
    require_contract: bool = True,
) -> IngestIdentity:
    file_path = Path(path)
    content = file_path.read_bytes()
    normalized_headers = {str(key).lower(): _text(value) for key, value in headers.items()}
    supplied = any(normalized_headers.get(key) for key in REQUIRED_HEADERS[1:])
    if not require_contract and not supplied:
        raise IngestContractError("legacy_upload", "未提供衝突保護 headers", status_code=0)
    missing = [key for key in REQUIRED_HEADERS if not normalized_headers.get(key)]
    if missing:
        raise IngestContractError("headers_missing", "缺少衝突保護 headers", fields=missing)

    metadata = read_km_metadata(file_path)
    missing_metadata = [field for field in REQUIRED_METADATA if not metadata.get(field)]
    if missing_metadata:
        raise IngestContractError("metadata_missing", "KM_Metadata 缺少必要欄位", fields=missing_metadata)
    # The multipart transport name may be an agent-generated envelope name
    # (for example ANRITSU__...__report.xlsx), so it is not required to equal
    # the report's originalFileName. Reject path-like metadata, but preserve
    # the producer's original report name as logical metadata.
    if Path(metadata["originalFileName"]).name != metadata["originalFileName"]:
        raise IngestContractError("metadata_invalid", "originalFileName 不可包含路徑", fields=["originalFileName"])

    actual_hash = hashlib.sha256(content).hexdigest()
    if metadata.get("ingestFileHash") and metadata["ingestFileHash"].lower() != actual_hash:
        raise IngestContractError("hash_mismatch", "ingestFileHash 與接收檔案內容不一致", fields=["ingestFileHash"])
    if not re.fullmatch(r"[0-9a-fA-F]{64}", metadata["sourceFileHash"]):
        raise IngestContractError("metadata_invalid", "sourceFileHash 必須是 SHA-256", fields=["sourceFileHash"])

    expected_doc = build_document_id(extraction_mode, metadata)
    expected_key = build_idempotency_key(metadata)
    comparisons = {
        "x-kb-source-system": metadata["sourceSystem"],
        "x-kb-environment-id": metadata["environmentId"],
        "x-kb-run-id": metadata["runId"],
        "x-kb-artifact-type": metadata["artifactType"],
        "x-kb-document-id": expected_doc,
        "idempotency-key": expected_key,
    }
    mismatches = [key for key, expected in comparisons.items() if normalized_headers.get(key) != expected]
    if mismatches or metadata["documentId"] != expected_doc or metadata["idempotencyKey"] != expected_key:
        raise IngestContractError("metadata_mismatch", "headers、KM_Metadata 或計算身份不一致", fields=mismatches)

    return IngestIdentity(
        source_system=metadata["sourceSystem"], environment_id=metadata["environmentId"], project_id=metadata["projectId"],
        run_id=metadata["runId"], artifact_type=metadata["artifactType"], report_schema=metadata["reportSchema"],
        original_file_name=metadata["originalFileName"], source_file_hash=metadata["sourceFileHash"],
        ingest_file_hash=actual_hash, document_id=expected_doc, idempotency_key=expected_key,
        generated_at=metadata["generatedAt"],
    )
