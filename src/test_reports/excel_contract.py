"""Versioned, deterministic Excel contract for external 4G/5G test reports."""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SUPPORTED_ENVIRONMENTS = {"anritsu", "amarisoft"}
ALLOWED_VERDICTS = {"pass", "fail", "error", "skipped"}
REQUIRED_SHEETS = {
    "Manifest",
    "RadioConfig",
    "TestCases",
    "Measurements",
    "Verdicts",
    "RawArtifacts",
}
REQUIRED_MANIFEST_FIELDS = {
    "schema_version",
    "run_id",
    "environment",
    "project_code",
    "dut_model",
    "started_at",
    "finished_at",
    "overall_verdict",
}
TABLE_REQUIRED_HEADERS = {
    "TestCases": {"case_id", "name", "status"},
    "Measurements": {"case_id", "metric", "value", "unit"},
    "Verdicts": {"case_id", "verdict", "reason"},
    "RawArtifacts": {"artifact_path", "sha256"},
}
SAFE_RUN_ID = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._:-]{0,127}$")


class ReportValidationError(ValueError):
    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__("; ".join(errors))


def _key(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _read_manifest(sheet) -> dict[str, str]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}
    first = [_key(cell) for cell in rows[0]]
    if "key" in first and "value" in first:
        key_idx, value_idx = first.index("key"), first.index("value")
        return {
            _key(row[key_idx]): _text(row[value_idx])
            for row in rows[1:]
            if len(row) > max(key_idx, value_idx) and _key(row[key_idx])
        }
    headers = first
    values = rows[1] if len(rows) > 1 else ()
    return {
        header: _text(values[index] if index < len(values) else "")
        for index, header in enumerate(headers)
        if header
    }


def _read_table(sheet) -> tuple[list[str], list[dict[str, Any]]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [_key(value) for value in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return headers, records


def parse_and_validate_report(path: str | Path, attachment_hashes: dict[str, str] | None = None) -> dict:
    report_path = Path(path)
    errors: list[str] = []
    if report_path.suffix.lower() != ".xlsx":
        raise ReportValidationError(["report 必須是 .xlsx 檔案"])
    try:
        workbook = load_workbook(report_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ReportValidationError([f"無法讀取 Excel: {exc}"]) from exc

    try:
        missing_sheets = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
        if missing_sheets:
            errors.append(f"缺少工作表: {', '.join(missing_sheets)}")
        if "Manifest" not in workbook.sheetnames:
            raise ReportValidationError(errors)

        manifest = _read_manifest(workbook["Manifest"])
        missing_fields = sorted(field for field in REQUIRED_MANIFEST_FIELDS if not manifest.get(field))
        if missing_fields:
            errors.append(f"Manifest 缺少必要欄位: {', '.join(missing_fields)}")

        environment = manifest.get("environment", "").lower()
        manifest["environment"] = environment
        if environment and environment not in SUPPORTED_ENVIRONMENTS:
            errors.append("environment 必須是 anritsu 或 amarisoft")
        if manifest.get("schema_version") and manifest["schema_version"] != "1.0":
            errors.append("目前只支援 schema_version 1.0")
        if manifest.get("run_id") and not SAFE_RUN_ID.fullmatch(manifest["run_id"]):
            errors.append("run_id 格式不合法")
        verdict = manifest.get("overall_verdict", "").lower()
        manifest["overall_verdict"] = verdict
        if verdict and verdict not in ALLOWED_VERDICTS:
            errors.append("overall_verdict 必須是 Pass/Fail/Error/Skipped")

        tables: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in sorted(REQUIRED_SHEETS - {"Manifest"}):
            if sheet_name not in workbook.sheetnames:
                continue
            headers, records = _read_table(workbook[sheet_name])
            required = TABLE_REQUIRED_HEADERS.get(sheet_name, set())
            missing_headers = sorted(required - set(headers))
            if missing_headers:
                errors.append(f"{sheet_name} 缺少欄位: {', '.join(missing_headers)}")
            tables[sheet_name] = records

        cases = tables.get("TestCases", [])
        case_ids = [_text(row.get("case_id")) for row in cases]
        if not cases:
            errors.append("TestCases 至少需要一筆資料")
        if any(not case_id for case_id in case_ids):
            errors.append("TestCases.case_id 不可為空")
        duplicate_ids = sorted({case_id for case_id in case_ids if case_ids.count(case_id) > 1})
        if duplicate_ids:
            errors.append(f"TestCases.case_id 重複: {', '.join(duplicate_ids)}")

        known_cases = set(case_ids)
        for sheet_name in ("Measurements", "Verdicts"):
            for row_number, row in enumerate(tables.get(sheet_name, []), start=2):
                case_id = _text(row.get("case_id"))
                if case_id not in known_cases:
                    errors.append(f"{sheet_name} 第 {row_number} 列引用未知 case_id: {case_id}")
        for row_number, row in enumerate(tables.get("Measurements", []), start=2):
            try:
                float(row.get("value"))
            except (TypeError, ValueError):
                errors.append(f"Measurements 第 {row_number} 列 value 必須是數值")
        for row_number, row in enumerate(tables.get("Verdicts", []), start=2):
            row_verdict = _text(row.get("verdict")).lower()
            row["verdict"] = row_verdict
            if row_verdict not in ALLOWED_VERDICTS:
                errors.append(f"Verdicts 第 {row_number} 列 verdict 不合法")

        provided_hashes = {Path(name).name: value.lower() for name, value in (attachment_hashes or {}).items()}
        for row_number, row in enumerate(tables.get("RawArtifacts", []), start=2):
            artifact_name = Path(_text(row.get("artifact_path"))).name
            expected_hash = _text(row.get("sha256")).lower()
            if not re.fullmatch(r"[0-9a-f]{64}", expected_hash):
                errors.append(f"RawArtifacts 第 {row_number} 列 sha256 格式不合法")
            elif provided_hashes.get(artifact_name) != expected_hash:
                errors.append(f"附件不存在或 hash 不符: {artifact_name}")

        if errors:
            raise ReportValidationError(errors)
        return {
            "manifest": manifest,
            "radio_config": tables.get("RadioConfig", []),
            "test_cases": cases,
            "measurements": tables.get("Measurements", []),
            "verdicts": tables.get("Verdicts", []),
            "raw_artifacts": tables.get("RawArtifacts", []),
        }
    finally:
        workbook.close()


def _md(value: Any) -> str:
    return _text(value).replace("|", "\\|").replace("\n", " ")


def _table(headers: list[str], rows: list[dict[str, Any]]) -> str:
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
    lines.extend("| " + " | ".join(_md(row.get(header.lower().replace(" ", "_"))) for header in headers) + " |" for row in rows)
    return "\n".join(lines)


def render_report_markdown(report: dict) -> str:
    manifest = report["manifest"]
    lines = [
        f"# 4G/5G Test Report — {manifest['run_id']}",
        "",
        "## Manifest",
        "",
        _table(["Key", "Value"], [{"key": key, "value": value} for key, value in manifest.items()]),
        "",
        "## Radio Configuration",
        "",
    ]
    radio_rows = report.get("radio_config", [])
    radio_headers = list(radio_rows[0].keys()) if radio_rows else ["key", "value", "unit"]
    lines.extend([_table([h.replace("_", " ").title() for h in radio_headers], radio_rows), ""])

    verdicts = {str(row.get("case_id")): row for row in report.get("verdicts", [])}
    measurements_by_case: dict[str, list[dict]] = {}
    for measurement in report.get("measurements", []):
        measurements_by_case.setdefault(str(measurement.get("case_id")), []).append(measurement)

    for index, case in enumerate(report.get("test_cases", []), start=1):
        case_id = str(case.get("case_id"))
        lines.extend([
            f"## Test Case {index}: {_md(case_id)} — {_md(case.get('name'))}",
            "",
            f"Environment: **{_md(manifest['environment'])}**  ",
            f"Project: **{_md(manifest['project_code'])}**  ",
            f"DUT: **{_md(manifest['dut_model'])}**  ",
            f"Status: **{_md(case.get('status'))}**",
            "",
            "### Measurements",
            "",
            _table(["Metric", "Value", "Unit", "Lower Limit", "Upper Limit"], measurements_by_case.get(case_id, [])),
            "",
        ])
        verdict_row = verdicts.get(case_id, {})
        lines.extend([
            "### Verdict",
            "",
            _table(["Case Id", "Verdict", "Reason"], [verdict_row] if verdict_row else []),
            "",
        ])

    lines.extend(["## Raw Artifacts", "", _table(["Artifact Path", "Sha256"], report.get("raw_artifacts", [])), ""])
    return "\n".join(lines)


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def report_summary_json(report: dict) -> str:
    return json.dumps(report, ensure_ascii=False, default=str)
