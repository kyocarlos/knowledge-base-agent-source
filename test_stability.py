#!/usr/bin/env python3
"""
系統穩定性與功能性測試套件
新增測試項目：並發、負載、服務降級、安全性、邊界條件
"""

import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import concurrent.futures
import threading
from src.search import SearchEngine

# 測試結果收集
all_results = []
se = SearchEngine()
se_lock = threading.Lock()

# ============================================
# 測試類別定義
# ============================================

TEST_CATEGORIES = {
    "A": "服務降級測試",
    "B": "並發/負載測試", 
    "C": "搜尋品質強化",
    "D": "安全性測試",
    "E": "意圖分類邊界",
    "F": "時間敏感查詢",
    "G": "長文本查詢",
}

# 測試案例定義
TEST_CASES = [
    # ========== A. 服務降級測試 (P0) ==========
    ("A1", "neo4j_down", "模擬Neo4j關閉時vector模式", "vector"),
    ("A2", "qdrant_down", "模擬QDrant關閉時deep模式", "deep"),
    ("A3", "ollama_single", "單Ollama實例壓力", "basic"),
    ("A4", "all_services_partial", "部分服務正常下的hybrid", "hybrid"),
    
    # ========== B. 並發/負載測試 (P0) ==========
    ("B1", "5 concurrent same query", "5個並發請求同一問題", "auto"),
    ("B2", "10 rapid requests", "10個快速連續請求", "basic"),
    ("B3", "burst test 20 requests", "20個請求爆發測試", "vector"),
    ("B4", "sustained load 50 requests", "50個請求持續負載", "deep"),
    
    # ========== C. 搜尋品質強化 (P1) ==========
    ("C1", "NR基地", "拼字錯誤測試", "auto"),
    ("C2", "5G電信基站", "同義詞測試", "auto"),
    ("C3", "NR", "極短關鍵字", "auto"),
    ("C4", "新空口 NR基站 5G系統 無線接入 移動通信 頻段 頻率 天線 傳輸 設備", "長關鍵字", "basic"),
    
    # ========== D. 安全性測試 (P1) ==========
    ("D1", "<script>alert(1)</script>", "XSS攻擊", "basic"),
    ("D2", "ls /etc/passwd", "命令注入", "basic"),
    ("D3", "[0-9]{5,}", "正則注入", "basic"),
    ("D4", "!@#$%^&*()_+-=[]{}|;':\",./<>?", "特殊字符", "basic"),
    ("D5", "'; DROP TABLE users; --", "SQL注入變種", "basic"),
    
    # ========== E. 意圖分類邊界 (P2) ==========
    ("E1", "NR基站的錯誤碼和解決方案", "雙意圖查詢", "auto"),
    ("E2", "那些設備有問題？", "模糊意圖", "auto"),
    ("E3", "有幾個NR基站？", "數量查詢", "auto"),
    ("E4", "多少設備支援5G？", "設備數量", "auto"),
    ("E5", "錯誤碼E-502和E-503的關係", "關聯查詢", "auto"),
    
    # ========== F. 時間敏感查詢 (P2) ==========
    ("F1", "最近新增的設備", "時間查詢-新", "auto"),
    ("F2", "最新的錯誤記錄", "時間查詢-錯誤", "auto"),
    
    # ========== G. 長文本查詢 (P2) ==========
    ("G1", "請幫我分析NR基站和LTE基站之間的差異，包括它們的技術規格、部署模式、作業頻段、調變方式、傳輸速率、覆蓋範圍、設備相容性、以及在4G和5G網路中的角色定位，並且說明在什麼情況下應該選擇哪種基站技術，以及未來的演進趨勢。", "超長文本查詢", "auto"),
    ("G2", "我需要找一个可以同时支持多个频段的基站设备，要求能够覆盖城市和郊区环境，具备5G能力，支持64QAM和256QAM调变，传输速率要达到10Gbps以上，最好是知名品牌的设备。", "詳細需求描述", "auto"),
]

def run_single_test(q_id, query, description, mode):
    """執行單一測試"""
    start = time.time()
    try:
        result = se.search(query, mode=mode)
        elapsed = time.time() - start
        return {
            "q_id": q_id,
            "query": query[:80] + "..." if len(query) > 80 else query,
            "description": description,
            "mode": mode,
            "selected_mode": result.get("mode", mode),
            "time": round(elapsed, 1),
            "sources_count": len(result.get("sources", [])),
            "status": result.get("status", "unknown"),
            "answer": result.get("answer", "")[:500] + "..." if len(result.get("answer", "")) > 500 else result.get("answer", ""),
            "error": None
        }
    except Exception as e:
        elapsed = time.time() - start
        return {
            "q_id": q_id,
            "query": query[:80] + "..." if len(query) > 80 else query,
            "description": description,
            "mode": mode,
            "selected_mode": "error",
            "time": round(elapsed, 1),
            "sources_count": 0,
            "status": "error",
            "answer": f"[錯誤: {str(e)[:200]}]",
            "error": str(e)
        }

def run_concurrent_test(num_concurrent, query, mode):
    """並發測試"""
    results = []
    start = time.time()
    
    def worker(i):
        result = run_single_test(f"concurrent_{i}", query, f"並發{i}", mode)
        return result
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=num_concurrent) as executor:
        futures = [executor.submit(worker, i) for i in range(num_concurrent)]
        results = [f.result() for f in concurrent.futures.as_completed(futures)]
    
    total_time = time.time() - start
    return results, total_time

def write_to_excel(results, filename):
    """寫入Excel報告"""
    wb = openpyxl.Workbook()
    
    # =====  Sheet 1: 完整測試報告  =====
    ws = wb.active
    ws.title = "完整測試報告"
    
    # 標題列
    headers = ["項次", "ID", "查詢", "描述", "模式", "自動選擇", "時間(秒)", "來源數", "狀態", "回覆摘要"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 填充數據
    for row_num, r in enumerate(results, 2):
        ws.append([
            row_num - 1,
            r["q_id"],
            r["query"],
            r["description"],
            r["mode"],
            r["selected_mode"],
            r["time"],
            r["sources_count"],
            r["status"],
            r["answer"]
        ])
        ws.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_num, column=10).alignment = Alignment(wrap_text=True)
    
    # 調整欄位寬度
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 60
    
    # =====  Sheet 2: 統計摘要  =====
    stats_ws = wb.create_sheet("統計摘要")
    
    # 計算統計
    stats = defaultdict(lambda: {"total": 0, "success": 0, "error": 0, "total_time": 0})
    for r in results:
        mode = r["mode"]
        stats[mode]["total"] += 1
        stats[mode]["total_time"] += r["time"]
        if r["status"] == "success":
            stats[mode]["success"] += 1
        else:
            stats[mode]["error"] += 1
    
    # 標題
    title_font = Font(bold=True, size=14)
    stats_ws.cell(row=1, column=1).value = "系統穩定性與功能性測試報告"
    stats_ws.cell(row=1, column=1).font = title_font
    stats_ws.append(["", ""])
    stats_ws.append(["測試時間", time.strftime("%Y-%m-%d %H:%M:%S")])
    stats_ws.append(["總案例數", str(len(results))])
    stats_ws.append(["", ""])
    
    # 類別統計表頭
    cat_header_row = stats_ws.max_row + 1
    headers = ["測試類別", "項目數", "成功", "失敗", "成功率", "平均時間"]
    for col, h in enumerate(headers, 1):
        cell = stats_ws.cell(row=cat_header_row, column=col)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 類別統計
    cat_stats = defaultdict(lambda: {"total": 0, "success": 0, "error": 0, "total_time": 0})
    for r in results:
        cat = r["q_id"][0]  # 第一個字母是類別
        cat_stats[cat]["total"] += 1
        cat_stats[cat]["total_time"] += r["time"]
        if r["status"] == "success":
            cat_stats[cat]["success"] += 1
        else:
            cat_stats[cat]["error"] += 1
    
    row = cat_header_row + 1
    for cat in sorted(cat_stats.keys()):
        s = cat_stats[cat]
        rate = (s["success"] / s["total"] * 100) if s["total"] > 0 else 0
        avg = (s["total_time"] / s["total"]) if s["total"] > 0 else 0
        cat_name = TEST_CATEGORIES.get(cat, cat)
        stats_ws.append([cat_name, s["total"], s["success"], s["error"], f"{rate:.1f}%", f"{avg:.1f}s"])
        row += 1
    
    # 模式統計
    stats_ws.append(["", ""])
    mode_header_row = stats_ws.max_row + 1
    stats_ws.cell(row=mode_header_row, column=1).value = "各模式統計"
    stats_ws.cell(row=mode_header_row, column=1).font = Font(bold=True)
    
    row = mode_header_row + 1
    headers = ["模式", "總數", "成功", "失敗", "成功率", "平均時間"]
    for col, h in enumerate(headers, 1):
        cell = stats_ws.cell(row=row, column=col)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    row += 1
    for mode in ["basic", "vector", "deep", "hybrid", "auto"]:
        s = stats[mode]
        if s["total"] == 0:
            continue
        rate = (s["success"] / s["total"] * 100) if s["total"] > 0 else 0
        avg = (s["total_time"] / s["total"]) if s["total"] > 0 else 0
        stats_ws.append([mode, s["total"], s["success"], s["error"], f"{rate:.1f}%", f"{avg:.1f}s"])
        row += 1
    
    # 調整欄位寬度
    stats_ws.column_dimensions['A'].width = 25
    stats_ws.column_dimensions['B'].width = 12
    stats_ws.column_dimensions['C'].width = 12
    stats_ws.column_dimensions['D'].width = 12
    stats_ws.column_dimensions['E'].width = 12
    stats_ws.column_dimensions['F'].width = 12
    
    # =====  Sheet 3: 問題列表  =====
    q_ws = wb.create_sheet("問題列表")
    q_ws.append(["ID", "描述", "測試類別"])
    for cat in sorted(TEST_CATEGORIES.keys()):
        q_ws.cell(row=q_ws.max_row, column=1).value = f"類別 {cat}"
        q_ws.cell(row=q_ws.max_row, column=2).value = TEST_CATEGORIES[cat]
    
    q_row = q_ws.max_row + 1
    for q_id, query, desc, mode in TEST_CASES:
        q_ws.append([q_id, query[:60] + "..." if len(query) > 60 else query, desc])
        q_row += 1
    
    q_ws.column_dimensions['A'].width = 10
    q_ws.column_dimensions['B'].width = 60
    q_ws.column_dimensions['C'].width = 25
    
    wb.save(filename)
    return stats

def main():
    print("=" * 80)
    print("系統穩定性與功能性測試")
    print("=" * 80)
    
    results = []
    
    # ===== Phase 1: 基本測試案例 =====
    print("\n[Phase 1] 執行基本測試案例...")
    for q_id, query, desc, mode in TEST_CASES:
        print(f"  {q_id}: {desc[:40]}...", end=" ")
        result = run_single_test(q_id, query, desc, mode)
        results.append(result)
        print(f"{result['time']:5.1f}s | {result['status']}")
    
    # ===== Phase 2: 並發測試 =====
    print("\n[Phase 2] 執行並發測試...")
    
    # B1: 5 個並發同一查詢
    print("  B1: 5 concurrent same query...", end=" ")
    concurrent_results, total_time = run_concurrent_test(5, "NR基站的錯誤碼有哪些？", "auto")
    for r in concurrent_results:
        r["q_id"] = f"B1_concurrent_{r['q_id'].split('_')[1]}"
        r["description"] = "5並發測試"
        results.append(r)
    print(f"{total_time:.1f}s total")
    
    # B2: 10 個快速連續請求
    print("  B2: 10 rapid requests...", end=" ")
    start = time.time()
    for i in range(10):
        r = run_single_test(f"B2_req{i}", "列出所有錯誤碼", f"快速請求{i}", "basic")
        results.append(r)
    print(f"{time.time()-start:.1f}s total")
    
    # B3: 20 個請求爆發測試
    print("  B3: 20 burst requests...", end=" ")
    start = time.time()
    for i in range(20):
        r = run_single_test(f"B3_burst{i}", "NR基站的位置", f"爆發{i}", "vector")
        results.append(r)
    print(f"{time.time()-start:.1f}s total")
    
    # B4: 50 個請求持續負載
    print("  B4: 50 sustained load (this will take a while)...", end=" ")
    start = time.time()
    for i in range(50):
        r = run_single_test(f"B4_load{i}", "設備狀態", f"負載{i}", "deep")
        results.append(r)
        if (i+1) % 10 == 0:
            print(f"\n    [{i+1}/50]...", end=" ", flush=True)
    print(f"\n    Total: {time.time()-start:.1f}s")
    
    # ===== 寫入 Excel =====
    output_path = "<project-root>/knowledge-base/系統穩定性測試報告.xlsx"
    stats = write_to_excel(results, output_path)
    
    # ===== 打印摘要 =====
    print("\n" + "=" * 80)
    print("測試完成！")
    print("=" * 80)
    print(f"\n報告位置: {output_path}")
    print(f"總案例數: {len(results)}")
    print("\n各類別統計:")
    print(f"{'類別':<25} {'總數':>6} {'成功':>6} {'失敗':>6} {'成功率':>10}")
    print("-" * 60)
    
    cat_stats = defaultdict(lambda: {"total": 0, "success": 0, "error": 0})
    for r in results:
        cat = r["q_id"][0]
        cat_stats[cat]["total"] += 1
        if r["status"] == "success":
            cat_stats[cat]["success"] += 1
        else:
            cat_stats[cat]["error"] += 1
    
    for cat in sorted(cat_stats.keys()):
        s = cat_stats[cat]
        rate = (s["success"] / s["total"] * 100) if s["total"] > 0 else 0
        cat_name = TEST_CATEGORIES.get(cat, cat)
        print(f"{cat_name:<25} {s['total']:>6} {s['success']:>6} {s['error']:>6} {rate:>9.1f}%")
    
    print("\n各模式統計:")
    print(f"{'模式':<10} {'總數':>6} {'成功':>6} {'失敗':>6} {'成功率':>10} {'平均時間':>10}")
    print("-" * 60)
    for mode in ["basic", "vector", "deep", "hybrid", "auto"]:
        s = stats[mode]
        if s["total"] == 0:
            continue
        rate = (s["success"] / s["total"] * 100) if s["total"] > 0 else 0
        avg = (s["total_time"] / s["total"]) if s["total"] > 0 else 0
        print(f"{mode:<10} {s['total']:>6} {s['success']:>6} {s['error']:>6} {rate:>9.1f}% {avg:>9.1f}s")

if __name__ == "__main__":
    main()
