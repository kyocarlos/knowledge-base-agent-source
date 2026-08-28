#!/usr/bin/env python3
"""Create one synthetic WP1 acceptance fixture with one authoritative run ID."""

from __future__ import annotations

import argparse
import hashlib
import json
import secrets
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from production_run_id_gate import check_unique_production_run_id

from openpyxl import Workbook, load_workbook

RUN_PREFIX = "TR-E2E-WP1-PROD-"


def new_run_id() -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return f"{RUN_PREFIX}{stamp}-{secrets.token_hex(4)}"


def build_fixture(output_dir: Path, run_id: str | None = None) -> dict[str, str]:
    authoritative_id = run_id or new_run_id()
    if not authoritative_id.startswith(RUN_PREFIX):
        raise ValueError(f"run_id must start with {RUN_PREFIX}")
    output_dir.mkdir(parents=True, exist_ok=True)
    attachment = output_dir / "synthetic-e2e-log.txt"
    attachment.write_text("synthetic WP1 acceptance artifact; no real user data\n", encoding="utf-8")
    attachment_hash = hashlib.sha256(attachment.read_bytes()).hexdigest()

    workbook = Workbook()
    manifest = workbook.active
    manifest.title = "Manifest"
    manifest.append(["key", "value"])
    for key, value in {
        "schema_version": "1.0",
        "run_id": authoritative_id,
        "test_run_id": authoritative_id,
        "environment": "anritsu",
        "project_code": "E2E-WP1-PRODUCTION-SYNTHETIC",
        "dut_model": "E2E-DUT-NONPRODUCTION",
        "started_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "finished_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "overall_verdict": "Pass",
    }.items():
        manifest.append([key, value])
    for name, headers, rows in [
        ("RadioConfig", ["key", "value", "unit"], [["profile", "synthetic", ""]]),
        ("TestCases", ["case_id", "name", "status"], [["E2E-TC-01", "Synthetic contract check", "completed"]]),
        ("Measurements", ["case_id", "metric", "value", "unit"], [["E2E-TC-01", "synthetic_score", 1, "count"]]),
        ("Verdicts", ["case_id", "verdict", "reason"], [["E2E-TC-01", "Pass", "synthetic fixture only"]]),
        ("RawArtifacts", ["artifact_path", "sha256"], [[attachment.name, attachment_hash]]),
    ]:
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    report = output_dir / f"{authoritative_id}.xlsx"
    workbook.save(report)
    return {"run_id": authoritative_id, "fixture": str(report), "attachment": str(attachment)}


def fixture_run_id(fixture: Path) -> str:
    workbook = load_workbook(fixture, read_only=True, data_only=True)
    try:
        rows = list(workbook["Manifest"].iter_rows(values_only=True))
    finally:
        workbook.close()
    values = {str(row[0]).strip(): str(row[1]).strip() for row in rows[1:] if len(row) >= 2 and row[0]}
    if values.get("run_id") != values.get("test_run_id"):
        raise ValueError("Manifest run_id and test_run_id must be identical")
    return values["run_id"]


def request_contract(run_id: str) -> dict[str, str]:
    return {
        "X-E2E-Test-Run-ID": run_id,
        "Idempotency-Key": run_id,
        "cleanup_test_run_id": run_id,
    }


def validate_request_contract(fixture: Path, request_run_id: str) -> dict[str, str]:
    manifest_id = fixture_run_id(fixture)
    if manifest_id != request_run_id:
        raise ValueError("request run_id must equal Manifest run_id")
    attachment = fixture.parent / "synthetic-e2e-log.txt"
    if not attachment.is_file():
        raise ValueError("required synthetic attachment is missing")
    workbook = load_workbook(fixture, read_only=True, data_only=True)
    try:
        if "RawArtifacts" not in workbook.sheetnames:
            raise ValueError("RawArtifacts sheet is missing")
        rows = list(workbook["RawArtifacts"].iter_rows(values_only=True))
    finally:
        workbook.close()
    artifact_hash = hashlib.sha256(attachment.read_bytes()).hexdigest()
    if not any(len(row) >= 2 and row[0] == attachment.name and row[1] == artifact_hash for row in rows[1:]):
        raise ValueError("required synthetic attachment hash does not match Manifest contract")
    return request_contract(manifest_id)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--run-id")
    parser.add_argument("--prior-production-evidence-root", type=Path)
    args = parser.parse_args()
    if args.prior_production_evidence_root and args.run_id:
        check_unique_production_run_id(args.run_id, args.prior_production_evidence_root.resolve())
    result = build_fixture(args.output_dir.resolve(), args.run_id)
    if fixture_run_id(Path(result["fixture"])) != result["run_id"]:
        raise RuntimeError("generated fixture run ID verification failed")
    print(json.dumps({**result, "request_contract": validate_request_contract(Path(result["fixture"]), result["run_id"]), "secrets_included": False}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
